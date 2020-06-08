import re
import glob
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Alignment, Side, NamedStyle
from collections import OrderedDict


def compare():
	
	
	#Cargar archivo de excel a la variable informe
	semana_actual='/Users/Ileam Montoya/Dropbox/Claro/Optimization - Daily/Dashboards/Dashboards/Dashboard_Fija_W16 - Guatemala - Honduras.xlsx'
	semana_pasada='/Users/Ileam Montoya/Dropbox/Claro/Optimization - Daily/Dashboards/Consolidados/Dashboard_Fija_W13 - Consolidado.xlsx'
	informe=openpyxl.load_workbook(semana_actual)
	informe = add_named_style(date_style,informe)
	consolidado=openpyxl.load_workbook(semana_pasada)

	get_normal_tabs(consolidado,informe)
	duplicate_int_tabs(informe)
	drifting(consolidado,informe,semana_pasada)
	get_CMTS_tabs(consolidado,informe)
	search_contraparte(informe)

	informe.save(semana_actual)


red = PatternFill(fill_type='solid',
						start_color='FF0000',
						end_color='FF0000')

thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

date_style = NamedStyle(name = 'fecha_extra', number_format = 'DD/MM/YYYY', border = thin_border)

def add_named_style(style,workbook):
	"""
	Funcion requerida para solucionar problemas de agregacion de estilo al libro.
	"""
	if style.name not in workbook.style_names:
		workbook.add_named_style(style)
	else:
		workbook._named_styles.pop(workbook.style_names.index(style.name))
		workbook.add_named_style(style)	
	return workbook

def search_contraparte(informe):
	
	tabs=['Interfaces Recurrentes > 95', 'Interfaces Recurrentes >70< 95']

	for tab in tabs:
		lista = {}
		acciones=informe[tab]
		data_row, data_column = 3, 3
		salir_data='{} - {}'.format(acciones.cell(row=data_row,column=3).value,acciones.cell(row=data_row,column=4).value)
		gua_hn = ['GT','HN','Guatemala','Honduras']
		while salir_data != 'None - None':
			if acciones.cell(row=data_row,column=2).value in gua_hn:
				equipo=acciones.cell(row=data_row,column=3).value
				ip=acciones.cell(row=data_row,column=5).value
				combined=[equipo,ip]
				lista.setdefault(salir_data,combined)
			data_row+=1
			salir_data='{} - {}'.format(acciones.cell(row=data_row,column=3).value,acciones.cell(row=data_row,column=4).value)


		data_row, data_column = 3, 3

		salir_raw='{} - {}'.format(acciones.cell(row=data_row,column=3).value,acciones.cell(row=data_row,column=4).value)
		while salir_raw != 'None - None':
			descripcion=acciones.cell(row=data_row,column=6).value
			for item in lista:
				if descripcion and acciones.cell(row=data_row,column=2).value in gua_hn:
					if lista[item][0] and lista[item][1]:
						if (lista[item][0] in descripcion or lista[item][1] in descripcion) and acciones.cell(row=data_row,column=3).value not in descripcion:
							acciones.cell(row=data_row,column=15).value="Contraparte de equipo {}".format(lista[item][0])
							acciones.cell(row=data_row,column=15).fill=red
			data_row+=1
			salir_raw='{} - {}'.format(acciones.cell(row=data_row,column=3).value,acciones.cell(row=data_row,column=4).value)

def get_normal_tabs(consolidado,informe):

	tabs=[('EQ LINK CMTS > 70',11), ('CMTS DHCP POOL >80', 11), ('IPPOOL > 90',9), ('DSLAMS > 80',11), ('MSAN > 80',11), ('CORE >80',13), ('MPLS P >80',14), ('MPLS PE >80',13), ('Isla de App > 80%',8), ('Equipos > 80%',14),
			('ServiceApp > 40',9), ('Enlaces > 40',10), ('Hubspoke > 80%',10), ('Interfaces Fotonico > 95',9), ('Interfaces Recurrentes > 95',10), ('Interfaces Recurrentes >70< 95',10)]
	for tab in tabs:
		print('Processing '+tab[0]+' for normal tabs')
		dict_acc=OrderedDict()
		#Carga Hoja con la informacion de las correctiones realizadas en semanas anteriores a la variable acciones
		acciones=consolidado[tab[0]]
		#Carga Hoja con informacion de la semana que se tiene que analizar a la variable raw
		raw=informe[tab[0]]

		col_equipo=3
		col_interfaz=4
		accion=tab[1]
		fecha=accion+1

		if tab[0] == 'Equipos > 80%':
			data_row, data_column = 4, 3
		else:
			data_row, data_column = 3, 3
		#Hay que copiar a partir de las filas que contienen el equipo y la interfaz para que cuadre con estas filas y columnas
		salir_data='{} - {}'.format(acciones.cell(row=data_row,column=col_equipo).value,acciones.cell(row=data_row,column=col_interfaz).value)
		while salir_data != 'None - None':
			#Crea entrada en diccionario usando la hoja de Lista Previa
			#Concatena equipo-interfaz como entrada del diccionario y le agrega un tuple con los valores de accion, fecha y localidad
			
			# gua_hn = ['GT','HN','Guatemala','Honduras', 'Nicaragua', 'NI', 'El Salvador', 'ESV', 'Costa Rica', 'CR','Panama']
			gua_hn = ['GT','HN','Guatemala','Honduras','GUATEMALA','HONDURAS']
			ingenieria_columns = ['Control ID','ROL','CLASIFICACION','ID DEL PROYECTO','CLASIFICACION INGENIERÍA','COMENTARIO INGENIERÍA','COMENTARIO PLANIFICACION','COMENTARIO PROYECTOS']
			if acciones.cell(row=data_row,column=2).value in gua_hn:
				accion_value = acciones.cell(row=data_row,column=accion).value
				fecha_value = acciones.cell(row=data_row,column=fecha).value
				if tab[0] in ['Interfaces Recurrentes > 95','Interfaces Recurrentes >70< 95','Interfaces Fotonico > 95']:
					column = 13
					ingenieria_data=[]
					for i, title in enumerate(ingenieria_columns):
						ingenieria_data.append(acciones.cell(row=data_row,column=column+i).value)
					dict_acc.setdefault(salir_data,tuple([accion_value,fecha_value]+ingenieria_data))
				else:
					dict_acc.setdefault(salir_data,(accion_value,fecha_value))

			data_row+=1
			salir_data='{} - {}'.format(acciones.cell(row=data_row,column=col_equipo).value,acciones.cell(row=data_row,column=col_interfaz).value)

		if tab[0] == 'Equipos > 80%':
			data_row, data_column = 4, 3
		else:
			data_row, data_column = 3, 3
		#Hay que copiar a partir de las filas que contienen el equipo y la interfaz para que cuadre con estas filas y columnas
		salir_raw='{} - {}'.format(raw.cell(row=data_row,column=col_equipo).value,raw.cell(row=data_row,column=col_interfaz).value)
		while salir_raw != 'None - None':
			#Busca la combinacion de equipo-interfaz en la hoja de los valores actuales
			#En caso de que la entrada se encuentre en el diccionario, se copian los valores de accion, fecha y localidad a la fila y columna adecuada de la hoja actual
			if salir_raw in dict_acc:
				if tab[0] in ['Interfaces Recurrentes > 95','Interfaces Recurrentes >70< 95','Interfaces Fotonico > 95']:
					column = 16					
					for i in range(8):
						raw.cell(row=data_row,column=column+i).value = dict_acc[salir_raw][i+2]

				raw.cell(row=data_row,column=accion).value = dict_acc[salir_raw][0]
				raw.cell(row=data_row,column=fecha).value = dict_acc[salir_raw][1]
				raw.cell(row=data_row,column=fecha).style = 'fecha_extra'

			data_row+=1
			salir_raw='{} - {}'.format(raw.cell(row=data_row,column=col_equipo).value,raw.cell(row=data_row,column=col_interfaz).value)

def get_CMTS_tabs(consolidado,informe):

	CMTS=[('CMTS PORTADORAS > 80',13), ('CMTS PORTADORAS > 30 < 79',13)]
	for tab in CMTS:
		print('Processing '+tab[0]+' for CMTS')
		dict_acc=OrderedDict()
		acciones=consolidado[tab[0]]
		raw=informe[tab[0]]
		data_row, data_column = 3, 3
		col_elemento= 3
		col_portadora= 5
		accion=tab[1]
		fecha=accion+1

		salir_data= '{}{}'.format(acciones.cell(row=data_row,column=col_elemento).value,acciones.cell(row=data_row,column=col_portadora).value.replace(' ',''))
		while salir_data != 'NoneNone':
			gua_hn = ['GT','HN','Guatemala','Honduras', 'Nicaragua', 'NI', 'El Salvador', 'ESV', 'Costa Rica', 'CR','Panama']
			# gua_hn = ['GT','HN','Guatemala','Honduras']
			if acciones.cell(row=data_row,column=2).value in gua_hn:
				dict_acc.setdefault(salir_data,(acciones.cell(row=data_row,column=accion).value,acciones.cell(row=data_row,column=fecha).value))

			data_row+=1
			if acciones.cell(row=data_row,column=col_portadora).value:
				salir_data= '{}{}'.format(acciones.cell(row=data_row,column=col_elemento).value,acciones.cell(row=data_row,column=col_portadora).value.replace(' ',''))
			else:
				salir_data= 'NoneNone'

		data_row, data_column = 3, 3
		salir_raw= '{}{}'.format(raw.cell(row=data_row,column=col_elemento).value,raw.cell(row=data_row,column=col_portadora).value.replace(' ',''))
		while salir_raw != 'NoneNone':
			if salir_raw in dict_acc:
				raw.cell(row=data_row,column=accion).value = dict_acc[salir_raw][0]
				raw.cell(row=data_row,column=fecha).value = dict_acc[salir_raw][1]
				raw.cell(row=data_row,column=fecha).style = 'fecha_extra'

			data_row+=1
			if raw.cell(row=data_row,column=col_portadora).value:
				salir_raw= '{}{}'.format(raw.cell(row=data_row,column=col_elemento).value,raw.cell(row=data_row,column=col_portadora).value.replace(' ',''))
			else:
				salir_raw='NoneNone'

def duplicate_int_tabs(informe):
	tabs=['Interfaces Recurrentes > 95', 'Interfaces Recurrentes >70< 95']


	#Cargar archivo de excel a la variable informe
	search_tabs=['EQ LINK CMTS > 70', 'DSLAMS > 80', 'MSAN > 80', 'CORE >80', 'MPLS P >80', 'MPLS PE >80', 'Isla de App > 80%', 'Equipos > 80%',
			'ServiceApp > 40', 'Enlaces > 40', 'Hubspoke > 80%']
	lista = {}
	for tab in search_tabs:
		acciones=informe[tab]
		data_row, data_column = 3, 3
		salir_data='{} - {}'.format(acciones.cell(row=data_row,column=3).value,acciones.cell(row=data_row,column=4).value)
		while salir_data != 'None - None':
			
			# gua_hn = ['GT','HN','Guatemala','Honduras', 'Nicaragua', 'NI', 'El Salvador', 'ESV', 'Costa Rica', 'CR']
			gua_hn = ['GT','HN','Guatemala','Honduras']
			if acciones.cell(row=data_row,column=2).value in gua_hn:
				lista.setdefault(salir_data,tab)
			data_row+=1
			salir_data='{} - {}'.format(acciones.cell(row=data_row,column=3).value,acciones.cell(row=data_row,column=4).value)

	for tab in tabs:
		print('Processing '+tab+' for duplicates')
		#Carga Hoja con informacion de la semana que se tiene que analizar a la variable raw
		raw=informe[tab]

		data_row, data_column = 3, 3
		#Hay que copiar a partir de las filas que contienen el equipo y la interfaz para que cuadre con estas filas y columnas
		salir_raw='{} - {}'.format(raw.cell(row=data_row,column=3).value,raw.cell(row=data_row,column=4).value)
		while salir_raw != 'None - None':
			#Busca la combinacion de equipo-interfaz en la hoja de los valores actuales
			#En caso de que la entrada se encuentre en el diccionario, se copian los valores de accion, fecha y localidad a la fila y columna adecuada de la hoja actual
			if salir_raw in lista:
				raw.cell(row=data_row,column=10).value = 'REPORTADO EN HOJA {}'.format(lista[salir_raw])
			data_row+=1
			salir_raw='{} - {}'.format(raw.cell(row=data_row,column=3).value,raw.cell(row=data_row,column=4).value)

def drifting(consolidado,informe,semana_pasada):
	tabs=[(('Interfaces Recurrentes > 95',10),('Interfaces Recurrentes >70< 95',10)),(('Interfaces Recurrentes >70< 95',10),('Interfaces Recurrentes > 95',10))]
	pasada=re.search('W(\d\d).+.xlsx',semana_pasada).group(1)
	for tab in tabs:
		print('Processing {} and {} for drifting'.format(tab[0][0],tab[1][0]))
		dict_acc=OrderedDict()
		#Carga Hoja con la informacion de las correctiones realizadas en semanas anteriores a la variable acciones
		acciones=consolidado[tab[0][0]]
		#Carga Hoja con informacion de la semana que se tiene que analizar a la variable raw
		raw=informe[tab[1][0]]

		col_equipo=3
		col_interfaz=4
		accion=tab[0][1]
		fecha=accion+1
		if tab == 'Equipos > 80%':
			data_row, data_column = 4, 3
		else:
			data_row, data_column = 3, 3
		#Hay que copiar a partir de las filas que contienen el equipo y la interfaz para que cuadre con estas filas y columnas
		salir_data='{} - {}'.format(acciones.cell(row=data_row,column=col_equipo).value,acciones.cell(row=data_row,column=col_interfaz).value)
		while salir_data != 'None - None':
			#Crea entrada en diccionario usando la hoja de Lista Previa
			#Concatena equipo-interfaz como entrada del diccionario y le agrega un tuple con los valores de accion, fecha y localidad
			
			# gua_hn = ['GT','HN','Guatemala','Honduras', 'Nicaragua', 'NI', 'El Salvador', 'ESV', 'Costa Rica', 'CR','Panama']
			gua_hn = ['GT','HN','Guatemala','Honduras']
			if acciones.cell(row=data_row,column=2).value in gua_hn:
				dict_acc.setdefault(salir_data,(acciones.cell(row=data_row,column=accion).value,acciones.cell(row=data_row,column=fecha).value))

			data_row+=1
			salir_data='{} - {}'.format(acciones.cell(row=data_row,column=col_equipo).value,acciones.cell(row=data_row,column=col_interfaz).value)

		if tab == 'Equipos > 80%':
			data_row, data_column = 4, 3
		else:
			data_row, data_column = 3, 3
		#Hay que copiar a partir de las filas que contienen el equipo y la interfaz para que cuadre con estas filas y columnas
		salir_raw='{} - {}'.format(raw.cell(row=data_row,column=col_equipo).value,raw.cell(row=data_row,column=col_interfaz).value)
		while salir_raw != 'None - None':
			#Busca la combinacion de equipo-interfaz en la hoja de los valores actuales
			#En caso de que la entrada se encuentre en el diccionario, se copian los valores de accion, fecha y localidad a la fila y columna adecuada de la hoja actual
			if salir_raw in dict_acc:
				raw.cell(row=data_row,column=accion+2).value = '{} de la Semana {}'.format(tab[0][0],pasada)
				raw.cell(row=data_row,column=accion+3).value = dict_acc[salir_raw][0]
				raw.cell(row=data_row,column=accion+4).value = dict_acc[salir_raw][1]

			data_row+=1
			salir_raw='{} - {}'.format(raw.cell(row=data_row,column=col_equipo).value,raw.cell(row=data_row,column=col_interfaz).value)

if __name__ == "__main__":

	compare()