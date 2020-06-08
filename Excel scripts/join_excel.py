import re, glob, os, openpyxl
from openpyxl.styles import Font, PatternFill, Border, Alignment, Side, NamedStyle
from collections import OrderedDict

red = PatternFill(fill_type='solid',
						start_color='FFFF00',
						end_color='FFFF00')
thin_border = Border(left=Side(style='thin'), 
					right=Side(style='thin'), 
					top=Side(style='thin'), 
					bottom=Side(style='thin'))
date_style = NamedStyle(name = 'fecha_extra', number_format = 'DD/MM/YYYY', border = thin_border, fill = red)

def add_named_style(style,workbook):
	"""
	Funcion utilizada para evitar error de agregacion de estilo al libro.


	USO EJEMPLO:
	name = NamedStyle(name = 'nombre', number_format = 'DD/MM/YYYY', border = thin_border)
	workbook = add_named_style(name,workbook)

	USO AL APLICAR EL ESTILO A UNA CELDA:
	sheet.cell(row = row, column = column).style = 'nombre'
	"""
	if style.name not in workbook.style_names:
		workbook.add_named_style(style)
	else:
		workbook._named_styles.pop(workbook.style_names.index(style.name))
		workbook.add_named_style(style)	
	return workbook

def walk_it():
	#NOSE INCLUYE ('IPPOOL > 90',12) procesar por separado
	tabs=[('EQ LINK CMTS > 70', 11), ('CMTS DHCP POOL >80', 11), ('IPPOOL > 90',9), ('DSLAMS > 80',11), ('MSAN > 80',11), ('CORE >80',13), ('MPLS P >80',14), ('MPLS PE >80',12), ('Isla de App > 80%',8), ('Equipos > 80%',14),
			('ServiceApp > 40',9), ('Enlaces > 40',10), ('Hubspoke > 80%',10), ('Interfaces Fotonico > 95',9), ('Interfaces Recurrentes > 95',10), ('Interfaces Recurrentes >70< 95',10)]
	eq_int=OrderedDict()
	current_week = 'Dashboard_Fija_W13 - Consolidado.xlsx'
	
	wb=openpyxl.load_workbook('Dashboard_Fija_W13 - Nicaragua, Costa Rica y El Salvador.xlsx')
	uni=openpyxl.load_workbook(current_week)
	uni = add_named_style(date_style,uni)
	for tab in tabs:
		eq_int=OrderedDict()
		print(tab[0])
		sheet=wb[tab[0]]
		if tab[0] == 'Equipos > 80%':
			row, column = 4, 3	
		else:
			row, column = 3, 3
		salir=sheet.cell(row=row,column=column).value
		while salir != None:
			eq_regex = re.compile(r'^([_A-Za-z0-9-]+)')
			equipo = eq_regex.search(sheet.cell(row=row,column=3).value).group()
			int_ip = sheet.cell(row=row,column=4).value
			accion = sheet.cell(row=row,column=tab[1]).value
			fecha = sheet.cell(row=row,column=tab[1]+1).value
			jerarquia = sheet.cell(row=row,column=tab[1]+2).value
			servicio = sheet.cell(row=row,column=tab[1]+3).value
			propagacion = sheet.cell(row=row,column=tab[1]+4).value

			interface=(equipo,int_ip,accion,fecha,jerarquia,servicio,propagacion)

			interface_text='{} - {}'.format(interface[0],interface[1])
			eq_int.setdefault(interface_text,interface)
			row+=1
			salir=sheet.cell(row=row,column=3).value
		
		sheet=uni[tab[0]]
		
		if tab[0] == 'Equipos > 80%':
			row, column = 4, 3
		else:
			row, column = 3, 3
		salir=sheet.cell(row=row,column=column).value
		while salir != None:
			eq_regex = re.compile(r'^([_A-Za-z0-9-]+)')
			equipo = eq_regex.search(sheet.cell(row=row,column=3).value).group()
			interface=(equipo,sheet.cell(row=row,column=4).value)
			interface_text='{} - {}'.format(interface[0],interface[1])
			if interface_text in eq_int and sheet.cell(row=row,column=2).value.strip() in ['El Salvador','ESV','Costa Rica','CR','Nicaragua','NI','EL_SALVADOR','NICARAGUA','COSTA_RICA','Panama','PANAMA']:
				sheet.cell(row=row,column=tab[1]).value = eq_int[interface_text][2]
				sheet.cell(row=row,column=tab[1]).fill=red
				sheet.cell(row=row,column=tab[1]+1).value = eq_int[interface_text][3]
				sheet.cell(row=row,column=tab[1]+1).style='fecha_extra'
				sheet.cell(row=row,column=tab[1]+2).value = eq_int[interface_text][4]
				sheet.cell(row=row,column=tab[1]+2).fill=red
			row+=1
			salir=sheet.cell(row=row,column=3).value

	uni.save(current_week)


if __name__ == "__main__":

	walk_it()